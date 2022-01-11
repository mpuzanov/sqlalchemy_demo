"""
xlsx (Excel 2007 и выше) - openpyxl, xlsxwriter
"""
from os import name
from typing import List
import openpyxl
from openpyxl.utils import get_column_letter
import datetime
from decimal import Decimal
from pydantic import BaseModel
from typing import Optional


class Column(BaseModel):
    field: str
    name: str
    format: Optional[str]
    width: Optional[str]
    func: Optional[str]
    num: Optional[int]


class Columns(BaseModel):
    fields: List[Column]


def query_to_excel2007_2(query_result,
                         file_name: str = 'file.xlsx',
                         sheet_name: str = 'Лист 1',
                         startrow=1,
                         columns=[]):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    if columns:
        columns = Columns(fields=columns)

    if not columns:  # если поля не заданы
        for key in query_result[0].keys():
            col = Column()
            col.field = col.name = key
            columns.append(col)

    row = startrow-1

    # добавляем заголовки в файл
    for col_index, val in enumerate(columns.fields):
        column_name = val.name
        ws.cell(row=row, column=col_index+1).value = column_name
    for cell in ws[row]:
        cell.style = "Headline 3"  # "header1" Total

    # ===================================================

    for item in query_result:
        row += 1
        for col, val in enumerate(columns.fields):
            columndata = item.get(val.field)

            if not val.format:
                # определяем формат поля для вывода
                style = 'General'
                if isinstance(columndata, (datetime.date, datetime.datetime)):
                    columndata = columndata.strftime(
                        "%d.%m.%Y")  # для расчета ширины колонки
                    style = 'dd-mm-yyyy'
                elif isinstance(columndata, Decimal):
                    style = '#,##0.#0'  # if columndata != 0 else 'General'
            else:
                style = val.format

            ws.cell(row=row, column=col+1, value=columndata)
            ws.cell(row=row, column=col+1).number_format = style

    # установка итогов
    for i, val in enumerate(columns.fields):
        col_name = get_column_letter(i+1)
        # print(f"{i=}, {val.func=} {col_name=} {startrow=} {row=}")
        if val.func:
            if val.func.lower() == 'sum':
                ws[f'{col_name}{row+1}'] = f"= SUM({col_name}{startrow}:{col_name}{row})"
                ws.cell(row=row+1, column=i+1).style = "Total"
                ws.cell(row=row+1, column=i+1).number_format = "#,##0.#0"

    for i, val in enumerate(columns.fields):
        col_name = get_column_letter(1 + i)
        ws.column_dimensions[col_name].width = val.width if val.width else 15

    wb.save(file_name)
