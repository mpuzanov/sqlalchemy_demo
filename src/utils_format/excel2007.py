"""
xlsx (Excel 2007 и выше) - openpyxl, xlsxwriter
"""
from os import name
import openpyxl
from openpyxl.utils import get_column_letter
import datetime
from decimal import Decimal


def query_to_excel2007(query_result,
                       file_name: str = 'file.xlsx',
                       sheet_name: str = 'Лист 1',
                       startrow=1,
                       columns={}):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    if not columns:  # если поля не заданы - создаем словарь колонок
        columns = {key: key for key in query_result[0].keys()}

    # header_list = query_result[0].keys()
    # ws.append([i for i in header_list])  # добавляем заголовки

    row = startrow-1

    # добавляем заголовки
    # for i in range(len(columns)):
    for col_index, col_name in enumerate(columns):
        column_data = columns.get(col_name)
        ws.cell(row=row, column=col_index+1).value = column_data
    for cell in ws[row]:
        cell.style = "Headline 3"  # "header1" Total

    # ===================================================

    for item in query_result:
        row += 1
        for col, col_name in enumerate(columns):
            columndata = item.get(col_name)

            # определяем формат поля для вывода
            style = 'General'
            if isinstance(columndata, (datetime.date, datetime.datetime)):
                columndata = columndata.strftime(
                    "%d.%m.%Y")  # для расчета ширины колонки
                style = 'dd-mm-yyyy'
            elif isinstance(columndata, Decimal):
                style = '#,##0.#0'  # if columndata != 0 else 'General'

            ws.cell(row=row, column=col+1, value=columndata)
            ws.cell(row=row, column=col+1).number_format = style

    adjust_column_width_from_col(ws, 2, 1, ws.max_column)

    wb.save(file_name)


def adjust_column_width_from_col(ws, min_row, min_col, max_col):
    column_widths = []
    min_width = 10

    for i, col in \
            enumerate(
                ws.iter_cols(min_col=min_col, max_col=max_col, min_row=min_row)
            ):

        for cell in col:
            value = cell.value
            if value is not None:

                if isinstance(value, str) is False:
                    value = str(value)

                try:
                    column_widths[i] = max(column_widths[i], len(value))
                except IndexError:
                    column_widths.append(len(value))

    for i, width in enumerate(column_widths):

        col_name = get_column_letter(min_col + i)
        adjusted_width = (column_widths[i] + 2) * 1.2
        if adjusted_width < min_width:
            adjusted_width = min_width
        ws.column_dimensions[col_name].width = adjusted_width
